"""Extract selected columns from the latest downloaded academic calendar."""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import openpyxl

DOWNLOAD_DIR = Path(__file__).with_name("downloads")
STATE_PATH = Path(__file__).with_name("state.json")
OUTPUT_PATH = Path(__file__).with_name("calendar_rows.json")
SHEET_NAME = "2025-2026 DERS-KAYIT TAKVİMİ"
TARGET_HEADERS = (
    "AKADEMİK DÖNEM",
    "KATEGORİ",
    "2025-2026 TARİH ARALIĞI",
)

TRANSLATION_TABLE = str.maketrans(
    {
        "ı": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ş": "s",
        "Ş": "s",
        "ç": "c",
        "Ç": "c",
        "ö": "o",
        "Ö": "o",
        "ü": "u",
        "Ü": "u",
    }
)

MONTHS_TR = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}

MONTH_LOOKUP = {
    "ocak": 1,
    "subat": 2,
    "mart": 3,
    "nisan": 4,
    "mayis": 5,
    "haziran": 6,
    "temmuz": 7,
    "agustos": 8,
    "eylul": 9,
    "ekim": 10,
    "kasim": 11,
    "aralik": 12,
}

RANGE_REGEX = re.compile(
    r"(?P<first>\d{1,2}(?:\s+\S+)?)\s*-\s*(?P<second>\d{1,2}(?:\s+\S+)?)$",
    flags=re.IGNORECASE,
)
DAY_MONTH_REGEX = re.compile(
    r"(?P<day>\d{1,2})(?:\s+(?P<month>\S+))?$",
    flags=re.IGNORECASE,
)
YEAR_REGEX = re.compile(r"(19|20)\d{2}$")


def load_state() -> Dict[str, Any]:
    if STATE_PATH.exists():
        with STATE_PATH.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    return {}


def load_latest_calendar() -> Path:
    state = load_state()
    saved_name = state.get("saved_filename")
    if saved_name:
        candidate = DOWNLOAD_DIR / saved_name
        if candidate.exists():
            return candidate

    try:
        return max(DOWNLOAD_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime)
    except ValueError as error:
        raise FileNotFoundError("No downloaded calendar file found.") from error


def sanitize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    return " ".join(str(value).split())


def format_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_display(value.date())
    if isinstance(value, date):
        month_name = MONTHS_TR.get(value.month, str(value.month))
        return f"{value.day:02d} {month_name} {value.year}"
    return sanitize_text(value)


def canonical_month(token: str) -> int:
    cleaned = token.translate(TRANSLATION_TABLE)
    cleaned = "".join(ch for ch in cleaned.lower() if ch.isalpha())
    if cleaned not in MONTH_LOOKUP:
        raise ValueError(f"Unknown month token: {token!r}")
    return MONTH_LOOKUP[cleaned]


def parse_day_month(part: str, fallback_month: int | None) -> Tuple[int, int]:
    match = DAY_MONTH_REGEX.fullmatch(part.strip())
    if not match:
        raise ValueError(f"Unsupported date segment: {part!r}")
    day = int(match.group("day"))
    month_token = match.group("month")
    if month_token is not None:
        month = canonical_month(month_token)
    else:
        if fallback_month is None:
            raise ValueError(f"Month information missing in segment: {part!r}")
        month = fallback_month
    return day, month


def parse_segment(segment: str, year: int, fallback_month: int | None) -> Tuple[List[Dict[str, str]], int]:
    compact = " ".join(segment.split())
    if not compact:
        return [], fallback_month

    range_match = RANGE_REGEX.fullmatch(compact)
    if range_match:
        first = range_match.group("first")
        second = range_match.group("second")
        second_day, second_month = parse_day_month(second, fallback_month)
        first_day, first_month = parse_day_month(first, second_month)

        start_year = year
        end_year = year
        if second_month < first_month:
            if first_month == 12 and second_month == 1:
                start_year = year - 1
            else:
                # Assume same calendar year if the order is unexpected by
                # falling back to the first month.
                second_month = first_month

        start_date = date(start_year, first_month, first_day)
        end_date = date(end_year, second_month, second_day)
        span = {"start": start_date.isoformat(), "end": end_date.isoformat()}
        return [span], second_month

    single_day, month = parse_day_month(compact, fallback_month)
    event_date = date(year, month, single_day)
    span = {"start": event_date.isoformat(), "end": event_date.isoformat()}
    return [span], month


def parse_tarih_string(display: str) -> Dict[str, Any]:
    spans: List[Dict[str, str]] = []
    normalized = display.replace("–", "-").replace("—", "-").strip()
    year_match = YEAR_REGEX.search(normalized)
    if not year_match:
        return {"raw": display, "spans": spans}

    year = int(year_match.group())
    prefix = normalized[: year_match.start()].strip()
    if not prefix:
        return {"raw": display, "spans": spans}

    segments = [segment.strip() for segment in prefix.split(",") if segment.strip()]
    parsed_segments: List[Tuple[List[Dict[str, str]], int]] = []
    next_month: int | None = None

    try:
        for segment in reversed(segments):
            segment_spans, month = parse_segment(segment, year, next_month)
            parsed_segments.append((segment_spans, month))
            next_month = month or next_month
    except ValueError:
        return {"raw": display, "spans": spans}

    for segment_spans, _ in reversed(parsed_segments):
        spans.extend(segment_spans)

    return {"raw": display, "spans": spans}


def parse_tarih_value(value: Any) -> Dict[str, Any]:
    display = format_display(value)
    if not display:
        return {"raw": "", "spans": []}

    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        iso_date = value.isoformat()
        return {"raw": display, "spans": [{"start": iso_date, "end": iso_date}]}

    return parse_tarih_string(display)


def build_column_map(header_row: Sequence[Any]) -> Dict[str, int]:
    column_map: Dict[str, int] = {}
    for index, header in enumerate(header_row):
        if header in TARGET_HEADERS:
            column_map[header] = index
    missing = [header for header in TARGET_HEADERS if header not in column_map]
    if missing:
        raise KeyError(f"Missing expected columns: {', '.join(missing)}")
    return column_map


def extract_rows(path: Path) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f"Sheet '{SHEET_NAME}' not found in workbook.")

    sheet = workbook[SHEET_NAME]
    column_map: Dict[str, int] | None = None
    results: List[Dict[str, Any]] = []

    for row in sheet.iter_rows(values_only=True):
        if column_map is None:
            try:
                column_map = build_column_map(row)
            except KeyError:
                continue
            else:
                continue

        akademik_donem = sanitize_text(row[column_map["AKADEMİK DÖNEM"]])
        kategori = sanitize_text(row[column_map["KATEGORİ"]])
        tarih_value = row[column_map["2025-2026 TARİH ARALIĞI"]]
        tarih = parse_tarih_value(tarih_value)

        if not (akademik_donem or kategori or tarih["raw"]):
            continue

        results.append(
            {
                "akademik_donem": akademik_donem,
                "kategori": kategori,
                "tarih": tarih,
            }
        )

    return results


def get_calendar_entries() -> List[Dict[str, Any]]:
    calendar_path = load_latest_calendar()
    return extract_rows(calendar_path)


def get_calendar_payload() -> Dict[str, Any]:
    calendar_path = load_latest_calendar()
    entries = extract_rows(calendar_path)
    state = load_state()
    return {
        "source": {
            "file_name": calendar_path.name,
            "downloaded_at": state.get("downloaded_at"),
            "url": state.get("last_url"),
        },
        "entries": entries,
    }


def save_payload(payload: Dict[str, Any]) -> None:
    with OUTPUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)


def main() -> None:
    payload = get_calendar_payload()
    save_payload(payload)
    print(
        "Saved",
        len(payload["entries"]),
        "rows to",
        OUTPUT_PATH,
    )


if __name__ == "__main__":
    main()
